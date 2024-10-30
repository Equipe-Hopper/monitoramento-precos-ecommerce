from botcity.web import WebBot, Browser, By
from webdriver_manager.chrome import ChromeDriverManager
from botcity.maestro import *
import time
import openpyxl
from openpyxl.chart import LineChart, Reference


BotMaestroSDK.RAISE_NOT_CONNECTED = False


def selecionar_produto(bot):
    while len(bot.find_elements('//*[@id="corePriceDisplay_desktop_feature_div"]/div[1]/span[2]/span[2]', By.XPATH)) < 1:
        bot.wait(1000)
        print('carrengado.')
    valor = bot.find_element('//*[@id="corePriceDisplay_desktop_feature_div"]/div[1]/span[2]/span[2]', By.XPATH).text
    valor_str=str(valor).replace('R$', "").splitlines()
    valor=valor_str[0]+ '.' +valor_str[1]
    # print(f"Valor contertido: {total}")
    return float(valor)

def acessar_amazon(bot):
    bot.browse("https://www.amazon.com.br/kindle-11geracao-preto/dp/B09SWTG9GF/ref=sr_1_1?__mk_pt_BR=%C3%85M%C3%85%C5%BD%C3%95%C3%91&crid=14Z1WC99NOCE0&dib=eyJ2IjoiMSJ9.YanHfJ7LSzbHFGfFLCYZLYnoU5r86AqxGWTKfKUMsSpK-sCv5GFYqzTjmlpVSnQQ6JFl_HpOA8f6mJlbkw51IdaS_XgiS6al-XedSjS1cl-p3NZfOXEWGm--BFUNWfW8FYZ8ZcDpNwB9PVG5BsGI7HXAYVHQg1c_af-EzcNxeqpA8qylYTo0b3Aa9pkyGAOewnWb0Mp3yZZJGQ830fcWl5gbvVhQd8oUQMT0oYa8fakzQnP97bweoYpoMSFq1HuVDyCL_5XRK9grxdjgy_DHyVHS4RJfHyLBySS17AqQgCM.GAax3C41BMlpvT3fJ-eSY4ouLxTEiG56kmB3_8l6ASQ&dib_tag=se&keywords=kindle+&qid=1726522381&sprefix=kinlde%2Caps%2C215&sr=8-1&ufe=app_do%3Aamzn1.fos.95de73c3-5dda-43a7-bd1f-63af03b14751")
    bot.wait(2000)

def datahoje():
    horario = time.localtime() 
    datahj= f'{horario[2]}/{horario[1]}/{horario[0]}'
    return datahj
    
def planilha(valor,data):

    wb=openpyxl.load_workbook('monitoramentoKindle.xlsx')
    aba_atual = wb.active
    aba_atual[f'A1'] = 'Valor'
    aba_atual[f'B1'] = 'Data'
    linha=aba_atual.max_row+1
    aba_atual[f'A{linha}'] = valor 
    aba_atual[f'B{linha}'] = data 

    wb.save("monitoramentoKindle.xlsx")

def gerar_grafico(arquivo_excel, posicao_grafico='E5'):
    # Carregar o arquivo Excel existente
    wb = openpyxl.load_workbook(arquivo_excel)
    aba_atual = wb.active

    # Verificar se há dados suficientes para o gráfico
    if aba_atual.max_row < 2:
        print("Não há dados suficientes para gerar um gráfico.")
        return

    # Criar um gráfico de linha
    chart = LineChart()
    chart.title = "Monitoramento de Preços do Kindle"
    chart.style = 13
    chart.x_axis.title = "Data"
    chart.y_axis.title = "Valor (R$)"
    
    # Selecionar dados para o gráfico (linhas da coluna A e B)
    dados = Reference(aba_atual, min_col=1, min_row=1, max_row=aba_atual.max_row+1)
    datas = Reference(aba_atual, min_col=2, min_row=2, max_row=aba_atual.max_row+1)
    
    # Adicionar dados ao gráfico
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(datas)
    
    # Adicionar o gráfico na planilha
    aba_atual.add_chart(chart, posicao_grafico)
    
    # Salvar a planilha com o gráfico
    wb.save(arquivo_excel)

def main():
    maestro = BotMaestroSDK.from_sys_args()
    execution = maestro.get_execution()

    print(f"Task ID is: {execution.task_id}")
    print(f"Task Parameters are: {execution.parameters}")

    bot = WebBot()
    bot.headless = False

    bot.browser = Browser.CHROME
    bot.driver_path = ChromeDriverManager().install()

    try:
        acessar_amazon(bot)
        valor=selecionar_produto(bot)
        data=datahoje()
        planilha(valor,data)
        print(valor)
        gerar_grafico('monitoramentoKindle.xlsx')
        
    except Exception as ex:
        print(ex)
        bot.save_screenshot('error.png')
    finally:
        print("Operação finalizada.")

  

def not_found(label):
    print(f"Element not found: {label}")


if __name__ == '__main__':
    main()